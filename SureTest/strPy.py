from openpyxl import load_workbook
from openpyxl.styles import Font, colors
import os


def getExcelData():
    excelPath = 'StrTestcase.xlsx'
    wb = load_workbook(filename = excelPath)
    ws = wb.active
    getNum1 = []
    getNum2 = []
    getOperator = []
    getExpectValue = []

    for loopidx in range(0 ,30):
        setIndex = 'C'+str(loopidx + 3)
        getNum1.append(ws[setIndex].value)

        setIndex = 'E' + str(loopidx + 3)
        getNum2.append(ws[setIndex].value)

        setIndex = 'D' + str(loopidx + 3)
        getOperator.append(ws[setIndex].value)

        setIndex = 'F' + str(loopidx + 3)
        getExpectValue.append(ws[setIndex].value)

    print('Data Parsing Finish...\n')
    wb.close()
    return (getNum1, getNum2, getOperator, getExpectValue)

def ExecuteCmd(getNum1, getNum2, getOperator):
    cmdPath = 'strTest.exe '
    resultList = []
    operList = ['?', '+', '-', '*']
    for loopidx in range(0, getOperator.__len__()):
        if getOperator[loopidx] in operList:
            if (getNum1[loopidx].find(' ') == -1) or (getNum2[loopidx].find(' ') == -1):
                setCmd = cmdPath + getNum1[loopidx] + ' ' + getOperator[loopidx] + ' ' + getNum2[loopidx]
                EqulText = os.popen(setCmd).read()
                findEqulText = EqulText.find('=')
                EqulText = EqulText[findEqulText + 1:]
                EqulText = EqulText.replace(' ', '')
                EqulText = EqulText.replace('\n', '')
                resultList.append(EqulText)
            else:
                resultList.append('ErrorCase')
        else:
            resultList.append('ErrorCase')

    return (resultList)

def ExpectToResult(resultList, getExpectValue):
    PassFailList = []
    for loopidx in range(0, resultList.__len__()):
        EqulText = resultList[loopidx]
        ExpectText = getExpectValue[loopidx]
        if EqulText == ExpectText:
            PassFailList.append('Pass')
        else:
            PassFailList.append('Fail')

    return (PassFailList)


def setExcelData(resultList, PassFailList):
    excelPath = 'StrTestcase.xlsx'
    wb = load_workbook(filename=excelPath)
    ws = wb.active
    for loopidx in range(0, resultList.__len__()):
        setIndex = 'G' + str(loopidx + 3)
        ws[setIndex] = resultList[loopidx]
        setIndex = 'H' + str(loopidx + 3)
        if PassFailList[loopidx] == 'Pass':
            ws[setIndex].font = Font(color=colors.BLACK)
        else:
            ws[setIndex].font = Font(color='FF0000')
        ws[setIndex] = PassFailList[loopidx]


    wb.save(filename=excelPath)
    wb.close()

if __name__ == '__main__':
    [getNum1, getNum2, getOperator, getExpectValue] = getExcelData()
    resultList = ExecuteCmd(getNum1, getNum2, getOperator)
    PassFailList = ExpectToResult(resultList, getExpectValue)
    setExcelData(resultList, PassFailList)